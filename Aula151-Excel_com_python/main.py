"""
http://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""

import openpyxl
from random import uniform

"""
pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_planilhas = pedidos.sheetnames  # <- Comando para ver o nome de quantas "abas" tem no excel.
planilha1 = pedidos['Página1']

print(planilha1['b4'].value, '\n')  # <- acessando coluna b-4

# for campo in planilha1['b']:
#     print(campo.value)

# for linha in planilha1['a1:c2']:
#     for coluna in linha:
#         print(coluna.value)

# for linha in planilha1:
#     print(linha[0].value, linha[1].value, linha[2].value, linha[3].value)

planilha1['B3'].value = 2200  # <- Alterando o valor da B3 na planilha.

for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha.xlsx')
"""
# Criando uma planilha (pagina de planilha dentro do mesmo arquivo):
planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']


for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'Luiz {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 2).value = f'Maria {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 3).value = f'Pereira {linha} {round(uniform(10, 100), 2)}'

planilha.save('nova_planilha_dois.xlsx')
